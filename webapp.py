import hmac
import json
import os
import shutil
import sqlite3
import subprocess
import tempfile
from datetime import datetime, timezone, timedelta
from html import escape
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from flask import (
    Flask,
    jsonify,
    redirect,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from runner import grade_assignment


app = Flask(__name__)

app.secret_key = os.environ.get(
    "FLASK_SECRET_KEY",
    "development-secret-change-this",
)

AUTOGRADER_ADMIN_PASSWORD = os.environ.get(
    "AUTOGRADER_ADMIN_PASSWORD"
)

PROJECT_ROOT = Path(__file__).resolve().parent
DATABASE_PATH = PROJECT_ROOT / "autograder.db"

ASSIGNMENTS = {
    "wk1ex1": PROJECT_ROOT / "examples" / "wk1ex1",
    "wk1ex3": PROJECT_ROOT / "examples" / "wk1ex3",
    "wk1ex4": PROJECT_ROOT / "examples" / "wk1ex4",

    "cs201lab1q1": (
        PROJECT_ROOT / "examples" / "cs201lab1q1"
    ),
    "cs201lab1q2": (
        PROJECT_ROOT / "examples" / "cs201lab1q2"
    ),
    "cs201lab1q3": (
        PROJECT_ROOT / "examples" / "cs201lab1q3"
    ),
}

ASSIGNMENT_COURSES = {
    "wk1ex1": "IS216-AY2627-T1",
    "wk1ex3": "IS216-AY2627-T1",
    "wk1ex4": "IS216-AY2627-T1",

    "cs201lab1q1": "CS201-AY2627-T1",
    "cs201lab1q2": "CS201-AY2627-T1",
    "cs201lab1q3": "CS201-AY2627-T1",
}

ASSIGNMENT_REPOSITORY_NAMES = {
    "wk1ex1": "wk1ex1-student-template",
    "wk1ex3": "wk1ex3-student-template",
    "wk1ex4": "wk1ex4-student-template",

    "cs201lab1q1": "cs201lab1q1-student-template",
    "cs201lab1q2": "cs201lab1q2-student-template",
    "cs201lab1q3": "cs201lab1q3-student-template",
}


def get_database():
    connection = sqlite3.connect(DATABASE_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def initialize_database():

    with get_database() as connection:

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                student_id TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                github_username TEXT NOT NULL UNIQUE,
                email TEXT,
                created_at TEXT NOT NULL
            )
            """
        )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS courses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                title TEXT NOT NULL,
                join_code TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            )
            """
        )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS enrolments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                course_id INTEGER NOT NULL,
                section TEXT,
                enrolled_at TEXT NOT NULL,

                UNIQUE(user_id, course_id),

                FOREIGN KEY(user_id)
                    REFERENCES users(id),

                FOREIGN KEY(course_id)
                    REFERENCES courses(id)
            )
            """
        )

        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS submissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                repository TEXT NOT NULL,
                assignment TEXT NOT NULL,
                commit_sha TEXT NOT NULL,
                score INTEGER NOT NULL,
                max_score INTEGER NOT NULL,
                passed INTEGER NOT NULL,
                tests_json TEXT NOT NULL,
                submitted_at TEXT NOT NULL,

                UNIQUE(
                    repository,
                    assignment,
                    commit_sha
                )
            )
            """
        )

        connection.execute(
            """
            INSERT OR IGNORE INTO courses (
                code,
                title,
                join_code,
                created_at
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                "IS216",
                "Web Application Development 2",
                "IS216-AY2627-T1",
                datetime.now(timezone.utc).isoformat(),
            ),
        )

        connection.execute(
            """
            INSERT OR IGNORE INTO courses (
                code,
                title,
                join_code,
                created_at
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                "CS201",
                "Data Structures and Algorithms",
                "CS201-AY2627-T1",
                datetime.now(timezone.utc).isoformat(),
            ),
        )

        submission_columns = connection.execute(
            "PRAGMA table_info(submissions)"
        ).fetchall()

        submission_column_names = {
            column[1]
            for column in submission_columns
        }

        if "course_id" not in submission_column_names:
            connection.execute(
                """
                ALTER TABLE submissions
                ADD COLUMN course_id INTEGER
                """
            )

        enrolment_columns = connection.execute(
            "PRAGMA table_info(enrolments)"
        ).fetchall()

        enrolment_column_names = {
            column[1]
            for column in enrolment_columns
        }

        if "section" not in enrolment_column_names:
            connection.execute(
                """
                ALTER TABLE enrolments
                ADD COLUMN section TEXT
                """
            )


def save_result(
    repository,
    assignment,
    course_id,
    commit_sha,
    score,
    max_score,
    passed,
    tests,
):
    submitted_at = datetime.now(timezone.utc).isoformat()

    with get_database() as connection:
        connection.execute(
            """
            INSERT INTO submissions (
                repository,
                assignment,
                course_id,
                commit_sha,
                score,
                max_score,
                passed,
                tests_json,
                submitted_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(repository, assignment, commit_sha)
            DO UPDATE SET
                course_id = excluded.course_id,
                score = excluded.score,
                max_score = excluded.max_score,
                passed = excluded.passed,
                tests_json = excluded.tests_json,
                submitted_at = excluded.submitted_at
            """,
            (
                repository,
                assignment,
                course_id,
                commit_sha,
                score,
                max_score,
                int(passed),
                json.dumps(tests),
                submitted_at,
            ),
        )


def format_section(section):

    if section is None:
        return "-"

    section = str(section).strip()

    if not section:
        return "-"

    if section.lower().startswith("g"):
        return section.upper()

    return f"G{section}"


def get_latest_submissions():

    with get_database() as connection:

        return connection.execute(
            """
            SELECT
                s.repository,
                s.assignment,
                s.commit_sha,
                s.score,
                s.max_score,
                s.passed,
                s.submitted_at,
                s.course_id,

                u.name AS student_name,
                u.student_id,
                u.email,
                u.github_username,

                e.section,
                c.code AS course_code,
                c.join_code AS course_join_code

            FROM submissions AS s

            LEFT JOIN users AS u
                ON LOWER(u.github_username) =
                   LOWER(
                       SUBSTR(
                           s.repository,
                           1,
                           INSTR(s.repository, '/') - 1
                       )
                   )

            LEFT JOIN enrolments AS e
                ON e.user_id = u.id
               AND e.course_id = s.course_id

            LEFT JOIN courses AS c
                ON c.id = s.course_id

            ORDER BY s.submitted_at DESC
            LIMIT 100
            """
        ).fetchall()


def format_excel_time(utc_string):

    if not utc_string:
        return ""

    utc_time = datetime.fromisoformat(utc_string)

    singapore_time = utc_time.astimezone(
        ZoneInfo("Asia/Singapore")
    )

    return singapore_time.strftime(
        "%d %b %Y %I:%M %p"
    )


def format_submission_time(utc_string):

    utc_time = datetime.fromisoformat(utc_string)

    sg_time = utc_time.astimezone(
        ZoneInfo("Asia/Singapore")
    )

    now = datetime.now(
        ZoneInfo("Asia/Singapore")
    )

    if sg_time.date() == now.date():
        return "Today " + sg_time.strftime("%I:%M %p")

    if sg_time.date() == (
        now.date() - timedelta(days=1)
    ):
        return "Yesterday " + sg_time.strftime("%I:%M %p")

    return sg_time.strftime(
        "%d %b %Y %I:%M %p"
    )


def admin_is_logged_in():
    return session.get("admin_logged_in") is True


def get_courses():
    with get_database() as connection:
        return connection.execute(
            """
            SELECT
                id,
                code,
                title,
                join_code,
                created_at
            FROM courses
            ORDER BY created_at DESC
            """
        ).fetchall()


def get_course_by_join_code(join_code):

    with get_database() as connection:
        return connection.execute(
            """
            SELECT
                id,
                code,
                title,
                join_code
            FROM courses
            WHERE join_code = ?
            """,
            (join_code,),
        ).fetchone()


def get_export_rows():

    with get_database() as connection:

        courses = connection.execute(
            """
            SELECT
                id,
                code,
                title,
                join_code
            FROM courses
            ORDER BY code, join_code
            """
        ).fetchall()

        export_rows = []

        for course in courses:

            students = connection.execute(
                """
                SELECT
                    users.id,
                    users.student_id,
                    users.name,
                    users.email,
                    users.github_username,
                    enrolments.section
                FROM users

                INNER JOIN enrolments
                    ON enrolments.user_id = users.id

                WHERE enrolments.course_id = ?

                ORDER BY
                    CAST(enrolments.section AS INTEGER),
                    users.name
                """,
                (course["id"],),
            ).fetchall()

            course_assignments = [
                assignment_key
                for assignment_key, join_code
                in ASSIGNMENT_COURSES.items()
                if join_code == course["join_code"]
            ]

            for student in students:

                for assignment_key in course_assignments:

                    repository_prefix = (
                        student["github_username"] + "/%"
                    )

                    submission = connection.execute(
                        """
                        SELECT
                            repository,
                            assignment,
                            commit_sha,
                            score,
                            max_score,
                            passed,
                            submitted_at
                        FROM submissions
                        WHERE course_id = ?
                          AND assignment = ?
                          AND LOWER(repository)
                              LIKE LOWER(?)
                        ORDER BY submitted_at DESC
                        LIMIT 1
                        """,
                        (
                            course["id"],
                            assignment_key,
                            repository_prefix,
                        ),
                    ).fetchone()

                    base_row = {
                        "course": course["code"],
                        "course_offering": course["join_code"],
                        "section": format_section(
                            student["section"]
                        ),
                        "student_id": student["student_id"],
                        "name": student["name"],
                        "email": student["email"] or "",
                        "github_username": (
                            student["github_username"]
                        ),
                        "assignment": assignment_key,
                    }

                    if submission:

                        status = (
                            "Passed"
                            if submission["passed"]
                            else "Failed"
                        )

                        export_rows.append({
                            **base_row,
                            "score": submission["score"],
                            "max_score": submission["max_score"],
                            "status": status,
                            "submitted_at": format_excel_time(
                                submission["submitted_at"]
                            ),
                            "commit_sha": submission["commit_sha"],
                        })

                    else:

                        export_rows.append({
                            **base_row,
                            "score": "",
                            "max_score": "",
                            "status": "Not submitted",
                            "submitted_at": "",
                            "commit_sha": "",
                        })

        return export_rows


@app.route("/join/<join_code>", methods=["GET", "POST"])
def join_course(join_code):

    course = get_course_by_join_code(join_code)

    if course is None:
        return "Course enrolment link not found.", 404

    error = None
    success = False

    if request.method == "POST":

        name = request.form.get(
            "name",
            "",
        ).strip()

        student_id = request.form.get(
            "student_id",
            "",
        ).strip()

        section = request.form.get(
            "section",
            "",
        ).strip()

        email = request.form.get(
            "email",
            "",
        ).strip().lower()

        github_username = request.form.get(
            "github_username",
            "",
        ).strip()

        if github_username.startswith("@"):
            github_username = github_username[1:]

        github_username = github_username.lower()

        if (
            not name
            or not student_id
            or not section
            or not email
            or not github_username
        ):
            error = "All fields are required."

        elif not section.isdigit() or int(section) < 1:
            error = (
                "Section must be a positive number "
                "(for example, 1 for G1)."
            )

        elif "@" not in email:
            error = "Please enter a valid email address."

        elif " " in github_username:
            error = "GitHub usernames cannot contain spaces."

        else:

            try:

                with get_database() as connection:

                    existing_user = connection.execute(
                        """
                        SELECT
                            id,
                            student_id,
                            name,
                            email,
                            github_username
                        FROM users
                        WHERE student_id = ?
                           OR LOWER(email) = LOWER(?)
                           OR LOWER(github_username) = LOWER(?)
                        """,
                        (
                            student_id,
                            email,
                            github_username,
                        ),
                    ).fetchone()

                    if existing_user is None:

                        cursor = connection.execute(
                            """
                            INSERT INTO users (
                                student_id,
                                name,
                                email,
                                github_username,
                                created_at
                            )
                            VALUES (?, ?, ?, ?, ?)
                            """,
                            (
                                student_id,
                                name,
                                email,
                                github_username,
                                datetime.now(
                                    timezone.utc
                                ).isoformat(),
                            ),
                        )

                        user_id = cursor.lastrowid

                    else:

                        same_student = (
                            existing_user["student_id"]
                            == student_id
                        )

                        existing_email = (
                            existing_user["email"] or ""
                        )

                        same_email = (
                            existing_email.lower()
                            == email.lower()
                        )

                        same_github = (
                            existing_user[
                                "github_username"
                            ].lower()
                            == github_username.lower()
                        )

                        if not (
                            same_student
                            and same_email
                            and same_github
                        ):
                            raise ValueError(
                                "The student ID, email, or GitHub "
                                "username is already registered "
                                "to another account."
                            )

                        user_id = existing_user["id"]

                        connection.execute(
                            """
                            UPDATE users
                            SET name = ?
                            WHERE id = ?
                            """,
                            (
                                name,
                                user_id,
                            ),
                        )

                    connection.execute(
                        """
                        INSERT INTO enrolments (
                            user_id,
                            course_id,
                            section,
                            enrolled_at
                        )
                        VALUES (?, ?, ?, ?)

                        ON CONFLICT(user_id, course_id)
                        DO UPDATE SET
                            section = excluded.section
                        """,
                        (
                            user_id,
                            course["id"],
                            section,
                            datetime.now(
                                timezone.utc
                            ).isoformat(),
                        ),
                    )

                success = True

            except ValueError as exception:
                error = str(exception)

            except sqlite3.IntegrityError:
                error = (
                    "That student ID, email, or GitHub username "
                    "has already been registered."
                )

    return render_template_string(
        """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1.0"
    >

    <title>Join {{ course["code"] }}</title>

    <style>
        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            padding: 40px 20px;
            font-family: Arial, sans-serif;
            background: #f4f6f8;
            color: #1f2937;
        }

        .card {
            max-width: 560px;
            margin: 30px auto;
            padding: 30px;
            background: white;
            border-radius: 12px;
            box-shadow: 0 3px 12px rgba(0, 0, 0, 0.08);
        }

        h1 {
            margin-bottom: 5px;
        }

        .course-title {
            margin-bottom: 6px;
            color: #6b7280;
        }

        .course-offering {
            margin-bottom: 25px;
            color: #6b7280;
            font-size: 0.92rem;
        }

        label {
            display: block;
            margin-bottom: 5px;
            font-weight: bold;
        }

        input {
            width: 100%;
            padding: 11px;
            margin-bottom: 17px;
            border: 1px solid #d1d5db;
            border-radius: 6px;
            font-size: 1rem;
        }

        button {
            width: 100%;
            padding: 12px;
            border: none;
            border-radius: 6px;
            background: #2563eb;
            color: white;
            font-weight: bold;
            font-size: 1rem;
            cursor: pointer;
        }

        button:hover {
            background: #1d4ed8;
        }

        .message {
            padding: 13px;
            margin-bottom: 20px;
            border-radius: 7px;
        }

        .error {
            color: #991b1b;
            background: #fee2e2;
        }

        .success {
            color: #166534;
            background: #dcfce7;
        }

        .hint {
            margin-top: -10px;
            margin-bottom: 17px;
            color: #6b7280;
            font-size: 0.9rem;
        }

        .summary {
            padding: 18px;
            background: #f9fafb;
            border-radius: 8px;
            line-height: 1.7;
        }
    </style>
</head>

<body>

    <div class="card">

        <h1>Join {{ course["code"] }}</h1>

        <div class="course-title">
            {{ course["title"] }}
        </div>

        <div class="course-offering">
            Course offering: {{ course["join_code"] }}
        </div>

        {% if success %}

            <div class="message success">
                Enrolment successful.
            </div>

            <div class="summary">
                <strong>
                    You have joined {{ course["code"] }}.
                </strong>

                <p>
                    Section:
                    <strong>
                        G{{ request.form.get("section", "") }}
                    </strong>
                </p>

                <p>
                    Your GitHub username is now linked to your
                    AutoGrade profile.
                </p>

                <p>
                    When you push an assignment to GitHub,
                    AutoGrade will associate the result with you.
                </p>
            </div>

        {% else %}

            {% if error %}
                <div class="message error">
                    {{ error }}
                </div>
            {% endif %}

            <form method="post">

                <label for="name">
                    Full name
                </label>

                <input
                    id="name"
                    name="name"
                    value="{{ request.form.get('name', '') }}"
                    required
                    autofocus
                >

                <label for="student_id">
                    Student ID
                </label>

                <input
                    id="student_id"
                    name="student_id"
                    value="{{ request.form.get('student_id', '') }}"
                    required
                >

                <label for="section">
                    Section
                </label>

                <input
                    id="section"
                    name="section"
                    type="number"
                    min="1"
                    step="1"
                    value="{{ request.form.get('section', '') }}"
                    placeholder="e.g. 1 for G1, 2 for G2"
                    required
                >

                <div class="hint">
                    Enter 1 for G1, 2 for G2, etc.
                </div>

                <label for="email">
                    Email address
                </label>

                <input
                    id="email"
                    name="email"
                    type="email"
                    value="{{ request.form.get('email', '') }}"
                    required
                >

                <div class="hint">
                    Enter the email address used in your
                    class records. It does not need to match
                    your GitHub email.
                </div>

                <label for="github_username">
                    GitHub username
                </label>

                <input
                    id="github_username"
                    name="github_username"
                    value="{{
                        request.form.get(
                            'github_username',
                            ''
                        )
                    }}"
                    placeholder="For example: cookiemo729"
                    required
                >

                <div class="hint">
                    Enter only your GitHub username, not the
                    full repository URL.
                </div>

                <button type="submit">
                    Join Course
                </button>

            </form>

        {% endif %}

    </div>

</body>
</html>
        """,
        course=course,
        error=error,
        success=success,
    )


@app.get("/")
def home():
    submissions = get_latest_submissions()

    rows = []

    for submission in submissions:

        passed = bool(submission["passed"])

        status_text = "Passed" if passed else "Failed"
        status_class = "passed" if passed else "failed"

        repository_owner = (
            submission["repository"].split("/")[0]
        )

        student = (
            submission["student_name"]
            or repository_owner
        )

        display_time = format_submission_time(
            submission["submitted_at"]
        )

        score_ratio = (
            submission["score"]
            / submission["max_score"]
            if submission["max_score"] > 0
            else 0
        )

        if score_ratio == 1:
            score_class = "score-full"
        elif score_ratio > 0:
            score_class = "score-partial"
        else:
            score_class = "score-zero"

        course_code = (
            submission["course_code"] or "-"
        )

        section_display = format_section(
            submission["section"]
        )

        rows.append(
            f"""
            <tr>
                <td>
                    <strong>{escape(str(student))}</strong>

                    <div class="student-github">
                        @{escape(repository_owner)}
                    </div>
                </td>

                <td>{escape(str(course_code))}</td>

                <td>{escape(section_display)}</td>

                <td>{escape(str(submission["assignment"]))}</td>

                <td>
                    <span class="score-badge {score_class}">
                        ●
                        {submission["score"]}/{submission["max_score"]}
                    </span>
                </td>

                <td>
                    <span class="status {status_class}">
                        {status_text}
                    </span>
                </td>

                <td>
                    <code>{escape(submission["commit_sha"][:7])}</code>
                </td>

                <td>{escape(display_time)}</td>
            </tr>
            """
        )

    if rows:
        table_body = "\n".join(rows)
    else:
        table_body = """
        <tr>
            <td colspan="8" class="empty">
                No submissions have been recorded yet.
            </td>
        </tr>
        """

    return f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1.0"
    >

    <title>AutoGrade Dashboard</title>

    <style>
        * {{
            box-sizing: border-box;
        }}

        body {{
            margin: 0;
            padding: 40px 20px;
            font-family: Arial, sans-serif;
            background: #f4f6f8;
            color: #1f2937;
        }}

        .container {{
            max-width: 1350px;
            margin: auto;
        }}

        h1 {{
            margin-bottom: 5px;
        }}

        .subtitle {{
            color: #6b7280;
            margin-bottom: 30px;
        }}

        .card {{
            background: white;
            border-radius: 12px;
            box-shadow: 0 3px 12px rgba(0, 0, 0, 0.08);
            overflow: hidden;
        }}

        .card-header {{
            padding: 20px 24px;
            border-bottom: 1px solid #e5e7eb;
        }}

        .student-github {{
            margin-top: 4px;
            color: #6b7280;
            font-size: 0.85rem;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
        }}

        th,
        td {{
            padding: 15px 18px;
            text-align: left;
            border-bottom: 1px solid #e5e7eb;
            white-space: nowrap;
        }}

        th {{
            background: #f9fafb;
            color: #4b5563;
        }}

        tr:last-child td {{
            border-bottom: none;
        }}

        .status {{
            display: inline-block;
            padding: 5px 11px;
            border-radius: 999px;
            font-size: 0.88rem;
            font-weight: bold;
        }}

        .passed {{
            background: #dcfce7;
            color: #166534;
        }}

        .failed {{
            background: #fee2e2;
            color: #991b1b;
        }}

        .score-badge {{
            display: inline-flex;
            align-items: center;
            gap: 7px;
            padding: 5px 11px;
            border-radius: 999px;
            font-weight: bold;
        }}

        .score-full {{
            background: #dcfce7;
            color: #166534;
        }}

        .score-partial {{
            background: #fef3c7;
            color: #92400e;
        }}

        .score-zero {{
            background: #fee2e2;
            color: #991b1b;
        }}

        .empty {{
            text-align: center;
            color: #6b7280;
            padding: 40px;
        }}

        code {{
            background: #f3f4f6;
            padding: 3px 6px;
            border-radius: 5px;
        }}

        .search-row {{
            display: flex;
            align-items: center;
            gap: 15px;
            margin-top: 14px;
            flex-wrap: wrap;
        }}

        .search-row input {{
            width: 480px;
            max-width: 100%;
            padding: 10px 12px;
            border: 1px solid #d1d5db;
            border-radius: 7px;
            font-size: 0.95rem;
        }}

        .search-row input:focus {{
            outline: none;
            border-color: #2563eb;
            box-shadow: 0 0 0 2px rgba(37, 99, 235, 0.12);
        }}

        #resultCount {{
            color: #6b7280;
            font-size: 0.9rem;
        }}

        @media (max-width: 900px) {{
            .card {{
                overflow-x: auto;
            }}

            table {{
                min-width: 1050px;
            }}
        }}
    </style>
</head>

<body>
    <div class="container">
        <h1>AutoGrade</h1>

        <div class="subtitle">
            Instructor submission dashboard
        </div>

        <div class="card">
            <div class="card-header">
                <strong>Latest submissions</strong>

                <div class="search-row">
                    <input
                        type="text"
                        id="submissionSearch"
                        placeholder="Search name, GitHub username, course, section, assignment..."
                        oninput="filterSubmissions()"
                    >

                    <span id="resultCount"></span>
                </div>
            </div>

            <table id="submissionsTable">
                <thead>
                    <tr>
                        <th>Student</th>
                        <th>Course</th>
                        <th>Section</th>
                        <th>Assignment</th>
                        <th>Score</th>
                        <th>Status</th>
                        <th>Commit</th>
                        <th>Submitted</th>
                    </tr>
                </thead>

                <tbody>
                    {table_body}
                </tbody>
            </table>
        </div>
    </div>

    <script>
        function filterSubmissions() {{
            const search = document
                .getElementById("submissionSearch")
                .value
                .toLowerCase()
                .trim();

            const terms = search
                .split(/\\s+/)
                .filter(Boolean);

            const table = document.getElementById(
                "submissionsTable"
            );

            const rows = table.querySelectorAll(
                "tbody tr"
            );

            let visible = 0;

            rows.forEach(row => {{
                const text = row.innerText.toLowerCase();

                const matches = terms.every(
                    term => text.includes(term)
                );

                row.style.display = matches ? "" : "none";

                if (matches) {{
                    visible++;
                }}
            }});

            document.getElementById(
                "resultCount"
            ).textContent =
                visible + " submission" +
                (visible === 1 ? "" : "s");
        }}

        filterSubmissions();
    </script>
</body>
</html>
"""


@app.route("/admin", methods=["GET", "POST"])
def admin():

    error = None

    if request.method == "POST":

        password = request.form.get("password", "")

        if (
            AUTOGRADER_ADMIN_PASSWORD
            and hmac.compare_digest(
                password,
                AUTOGRADER_ADMIN_PASSWORD,
            )
        ):
            session["admin_logged_in"] = True

            return redirect(
                url_for("admin")
            )

        error = "Incorrect password."

    if not admin_is_logged_in():

        return render_template_string(
            """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1.0"
    >

    <title>AutoGrade Admin</title>

    <style>
        body {
            margin: 0;
            padding: 40px 20px;
            font-family: Arial, sans-serif;
            background: #f4f6f8;
            color: #1f2937;
        }

        .card {
            max-width: 420px;
            margin: 80px auto;
            padding: 30px;
            background: white;
            border-radius: 12px;
            box-shadow: 0 3px 12px rgba(0, 0, 0, 0.08);
        }

        input {
            width: 100%;
            padding: 11px;
            margin: 8px 0 16px;
            box-sizing: border-box;
        }

        button {
            width: 100%;
            padding: 11px;
            border: none;
            border-radius: 6px;
            background: #2563eb;
            color: white;
            font-weight: bold;
            cursor: pointer;
        }

        .error {
            color: #991b1b;
            background: #fee2e2;
            padding: 10px;
            border-radius: 6px;
            margin-bottom: 15px;
        }
    </style>
</head>

<body>
    <div class="card">
        <h1>AutoGrade Admin</h1>

        {% if error %}
            <div class="error">
                {{ error }}
            </div>
        {% endif %}

        <form method="post">
            <label for="password">
                Admin password
            </label>

            <input
                id="password"
                name="password"
                type="password"
                required
                autofocus
            >

            <button type="submit">
                Sign in
            </button>
        </form>
    </div>
</body>
</html>
            """,
            error=error,
        )

    courses = get_courses()

    return render_template_string(
        """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">

    <meta
        name="viewport"
        content="width=device-width, initial-scale=1.0"
    >

    <title>AutoGrade Admin</title>

    <style>
        * {
            box-sizing: border-box;
        }

        body {
            margin: 0;
            padding: 40px 20px;
            font-family: Arial, sans-serif;
            background: #f4f6f8;
            color: #1f2937;
        }

        .container {
            max-width: 950px;
            margin: auto;
        }

        .topbar {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 25px;
        }

        .card {
            background: white;
            padding: 24px;
            border-radius: 12px;
            box-shadow: 0 3px 12px rgba(0, 0, 0, 0.08);
            margin-bottom: 24px;
        }

        input {
            width: 100%;
            padding: 10px;
            margin: 6px 0 14px;
        }

        button {
            padding: 10px 16px;
            border: none;
            border-radius: 6px;
            background: #2563eb;
            color: white;
            font-weight: bold;
            cursor: pointer;
        }

        .delete-button {
            background: #dc2626;
        }

        .delete-button:hover {
            background: #b91c1c;
        }

        table {
            width: 100%;
            border-collapse: collapse;
        }

        th,
        td {
            padding: 12px;
            text-align: left;
            border-bottom: 1px solid #e5e7eb;
        }

        th {
            background: #f9fafb;
        }

        a {
            color: #2563eb;
        }

        code {
            background: #f3f4f6;
            padding: 3px 6px;
            border-radius: 4px;
        }

        .export-button {
            display: inline-block;
            padding: 10px 16px;
            border-radius: 6px;
            background: #15803d;
            color: white;
            text-decoration: none;
            font-weight: bold;
            margin-bottom: 18px;
        }

        .export-button:hover {
            background: #166534;
        }
    </style>
</head>

<body>
    <div class="container">

        <a
            href="{{ url_for('admin_export_results') }}"
            class="export-button"
        >
            Export All Results
        </a>

        <div class="topbar">
            <div>
                <h1>AutoGrade Admin</h1>
                <p>Create and manage courses.</p>
            </div>

            <a href="{{ url_for('admin_logout') }}">
                Sign out
            </a>
        </div>

        <div class="card">
            <h2>Create course</h2>

            <form
                method="post"
                action="{{ url_for('admin_create_course') }}"
            >
                <label for="code">
                    Course code
                </label>

                <input
                    id="code"
                    name="code"
                    placeholder="IS216"
                    required
                >

                <label for="title">
                    Course title
                </label>

                <input
                    id="title"
                    name="title"
                    placeholder="Web Application Development 2"
                    required
                >

                <label for="join_code">
                    Join code
                </label>

                <input
                    id="join_code"
                    name="join_code"
                    placeholder="IS216-AY2627-T1"
                    required
                >

                <button type="submit">
                    Create course
                </button>
            </form>
        </div>

        <div class="card">
            <h2>Courses</h2>

            <table>
                <thead>
                    <tr>
                        <th>Code</th>
                        <th>Title</th>
                        <th>Join link</th>
                        <th>Action</th>
                    </tr>
                </thead>

                <tbody>
                    {% for course in courses %}
                        <tr>
                            <td>{{ course["code"] }}</td>

                            <td>{{ course["title"] }}</td>

                            <td>
                                <code>
                                    /join/{{ course["join_code"] }}
                                </code>
                            </td>

                            <td>
                                <form
                                    method="post"
                                    action="{{ url_for(
                                        'admin_delete_course',
                                        course_id=course['id']
                                    ) }}"
                                    onsubmit="
                                        return confirm(
                                            'Delete this course, its enrolments, and '
                                            + 'its submission records? This cannot be undone.'
                                        );
                                    "
                                >
                                    <button
                                        type="submit"
                                        class="delete-button"
                                    >
                                        Delete
                                    </button>
                                </form>
                            </td>
                        </tr>
                    {% else %}
                        <tr>
                            <td colspan="4">
                                No courses created.
                            </td>
                        </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>

    </div>
</body>
</html>
        """,
        courses=courses,
    )


@app.post("/admin/courses")
def admin_create_course():

    if not admin_is_logged_in():
        return redirect(
            url_for("admin")
        )

    code = request.form.get(
        "code",
        "",
    ).strip()

    title = request.form.get(
        "title",
        "",
    ).strip()

    join_code = request.form.get(
        "join_code",
        "",
    ).strip()

    if not code or not title or not join_code:
        return "All course fields are required.", 400

    try:
        with get_database() as connection:
            connection.execute(
                """
                INSERT INTO courses (
                    code,
                    title,
                    join_code,
                    created_at
                )
                VALUES (?, ?, ?, ?)
                """,
                (
                    code,
                    title,
                    join_code,
                    datetime.now(
                        timezone.utc
                    ).isoformat(),
                ),
            )

    except sqlite3.IntegrityError:
        return (
            "A course with that join code already exists.",
            409,
        )

    return redirect(
        url_for("admin")
    )


@app.post("/admin/courses/<int:course_id>/delete")
def admin_delete_course(course_id):

    if not admin_is_logged_in():
        return redirect(
            url_for("admin")
        )

    with get_database() as connection:

        course = connection.execute(
            """
            SELECT id, code, title
            FROM courses
            WHERE id = ?
            """,
            (course_id,),
        ).fetchone()

        if course is None:
            return "Course not found.", 404

        latest_submission = connection.execute(
            """
            SELECT MAX(submitted_at)
            FROM submissions
            WHERE course_id = ?
            """,
            (course_id,),
        ).fetchone()[0]

        if latest_submission:

            latest_time = datetime.fromisoformat(
                latest_submission
            )

            cutoff_time = datetime.now(
                timezone.utc
            ) - timedelta(days=365 * 3)

            if latest_time > cutoff_time:
                return (
                    "This course cannot be deleted because "
                    "its latest submission was less than "
                    "three years ago.",
                    409,
                )

        connection.execute(
            """
            DELETE FROM enrolments
            WHERE course_id = ?
            """,
            (course_id,),
        )

        connection.execute(
            """
            DELETE FROM submissions
            WHERE course_id = ?
            """,
            (course_id,),
        )

        connection.execute(
            """
            DELETE FROM courses
            WHERE id = ?
            """,
            (course_id,),
        )

    return redirect(
        url_for("admin")
    )


@app.get("/admin/export/results.xlsx")
def admin_export_results():

    if not admin_is_logged_in():
        return redirect(
            url_for("admin")
        )

    rows = get_export_rows()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "All Results"

    headers = [
        "Course",
        "Course Offering",
        "Section",
        "Student ID",
        "Name",
        "Email",
        "GitHub Username",
        "Assignment",
        "Score",
        "Max Score",
        "Status",
        "Submitted (SGT)",
        "Commit SHA",
    ]

    worksheet.append(headers)

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in rows:
        worksheet.append([
            row["course"],
            row["course_offering"],
            row["section"],
            row["student_id"],
            row["name"],
            row["email"],
            row["github_username"],
            row["assignment"],
            row["score"],
            row["max_score"],
            row["status"],
            row["submitted_at"],
            row["commit_sha"],
        ])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    column_widths = {
        "A": 12,
        "B": 22,
        "C": 10,
        "D": 15,
        "E": 24,
        "F": 30,
        "G": 22,
        "H": 18,
        "I": 10,
        "J": 12,
        "K": 16,
        "L": 24,
        "M": 42,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = (
        "autograde_results_"
        + datetime.now(
            ZoneInfo("Asia/Singapore")
        ).strftime("%Y%m%d_%H%M")
        + ".xlsx"
    )

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )


@app.get("/admin/logout")
def admin_logout():

    session.clear()

    return redirect(
        url_for("admin")
    )


@app.get("/health")
def health():
    return jsonify({"status": "ok"})


@app.post("/api/grade")
def grade():

    data = request.get_json(silent=True) or {}

    assignment_key = data.get("assignment")
    repository_name = data.get("repository")
    commit_sha = data.get("commit_sha")

    if not assignment_key:
        return jsonify({"error": "Missing assignment"}), 400

    if not repository_name:
        return jsonify({"error": "Missing repository"}), 400

    if not commit_sha:
        return jsonify({"error": "Missing commit_sha"}), 400

    assignment_path = ASSIGNMENTS.get(assignment_key)

    if assignment_path is None:
        return jsonify({"error": "Unknown assignment"}), 400

    course_join_code = ASSIGNMENT_COURSES.get(
        assignment_key
    )

    if not course_join_code:
        return jsonify({
            "error": "Assignment is not linked to a course"
        }), 500

    with get_database() as connection:

        course = connection.execute(
            """
            SELECT id, code, title, join_code
            FROM courses
            WHERE join_code = ?
            """,
            (course_join_code,),
        ).fetchone()

    if course is None:
        return jsonify({
            "error": "Course not found"
        }), 500

    course_id = course["id"]

    if "/" not in repository_name:
        return jsonify({
            "error": "Invalid repository name"
        }), 400

    repository_owner, repository_short_name = (
        repository_name.split("/", 1)
    )

    expected_repository_name = (
        ASSIGNMENT_REPOSITORY_NAMES.get(
            assignment_key
        )
    )

    if expected_repository_name is None:
        return jsonify({
            "error": "Unknown assignment repository"
        }), 400

    if repository_short_name != expected_repository_name:
        return jsonify({
            "error": (
                "Repository name does not match "
                "the assignment"
            )
        }), 403

    with get_database() as connection:

        student = connection.execute(
            """
            SELECT
                users.id,
                users.name,
                users.github_username,
                enrolments.section
            FROM users

            INNER JOIN enrolments
                ON enrolments.user_id = users.id

            WHERE
                LOWER(users.github_username)
                    = LOWER(?)

                AND enrolments.course_id = ?
            """,
            (
                repository_owner,
                course_id,
            ),
        ).fetchone()

        if student is None:
            return jsonify({
                "error": (
                    "GitHub account is not enrolled "
                    f"in {course['code']} AutoGrade"
                )
            }), 403

        if not student["section"]:
            return jsonify({
                "error": (
                    "Your course enrolment does not have a "
                    "section. Please re-enrol using the course "
                    "join page and enter your section."
                )
            }), 403

    repository_url = (
        f"https://github.com/{repository_name}.git"
    )

    work_folder = Path(
        tempfile.mkdtemp(prefix="autograde-")
    )

    submission_folder = work_folder / "submission"

    try:
        clone_result = subprocess.run(
            [
                "git",
                "clone",
                "--no-checkout",
                repository_url,
                str(submission_folder),
            ],
            capture_output=True,
            text=True,
            timeout=60,
        )

        if clone_result.returncode != 0:
            return jsonify({
                "error": "Unable to clone repository",
                "details": clone_result.stderr,
            }), 400

        checkout_result = subprocess.run(
            [
                "git",
                "-C",
                str(submission_folder),
                "checkout",
                "--detach",
                commit_sha,
            ],
            capture_output=True,
            text=True,
            timeout=30,
        )

        if checkout_result.returncode != 0:
            return jsonify({
                "error": "Unable to check out commit",
                "details": checkout_result.stderr,
            }), 400

        assignment, result = grade_assignment(
            str(assignment_path),
            str(submission_folder),
        )

        tests = [
            {
                "name": test.name,
                "passed": test.passed,
                "points": test.points,
                "awarded": (
                    test.points if test.passed else 0
                ),
            }
            for test in result.tests
        ]

        passed = result.score == result.max_score

        save_result(
            repository=repository_name,
            assignment=assignment_key,
            course_id=course_id,
            commit_sha=commit_sha,
            score=result.score,
            max_score=result.max_score,
            passed=passed,
            tests=tests,
        )

        return jsonify({
            "assignment": assignment_key,
            "title": assignment["title"],
            "repository": repository_name,
            "commit_sha": commit_sha,
            "score": result.score,
            "max_score": result.max_score,
            "passed": passed,
            "tests": tests,
        }), 200

    except subprocess.TimeoutExpired:
        return jsonify({
            "error": "Repository operation timed out"
        }), 408

    except Exception as error:
        app.logger.exception("Grading failed")

        return jsonify({
            "error": "Grading failed",
            "details": str(error),
        }), 500

    finally:
        shutil.rmtree(
            work_folder,
            ignore_errors=True,
        )


if __name__ == "__main__":
    initialize_database()

    app.run(
        host="127.0.0.1",
        port=8000,
    )
